# Data manipulation library; used to read Excel files into DataFrames and perform row/column operations
import pandas as pd

# Standard library module for writing and running asynchronous concurrent code
import asyncio

# Standard library to perform operations with files and folders
import os

# Standard library module to perform operations with time
import time

# Third-party library for retrieving information on running processes and open files
import psutil

# Low-level Excel (.xlsx) reader/writer; used to open and update individual cells in an existing workbook
from openpyxl import load_workbook

# Local helper module; provides IP Addressing operations for validaton, calculation, ...
import ip_addressing_ops

# Local helper module; provides API methods to retrieve and push data of network devices
import device_api_ops

# Local helper module; provides CLI methods to retrieve and push data of network devices
import device_cli_ops

# Local helper module; provides CLI SCP methods to transfer the files for the update
import device_file_transfer_ops

# Local helper module; provides utility functions such as get_absolute_path()
import common_helper_functions

# Standard library module for writing structured log messages (INFO, ERROR, etc.)
import logging

# Logging handler that auto-rotates the log file when it reaches a defined size limit
from logging.handlers import RotatingFileHandler

###################################################################################################################################

# SETUP LOGGING
# -------------
logger = logging.getLogger(__name__) # use the module's name as the name in the logs
logger.setLevel(logging.INFO) # set the logging level
log_file_path = 'execution_logs.log' # define the logging file path

# Use RotatingFileHandler
# log_file_path = the absolute path for the log file
# maxBytes: 5 * 1024
# backupCount=0: When the file is full, delete it and start a new one
handler = RotatingFileHandler(
    log_file_path, maxBytes=5*1024*1024, backupCount=0
)

# Set the format of the log messages
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)

# Add the formatted HANDLER to the logger
logger.addHandler(handler)

###################################################################################################################################

# CREATE DEVICES DATAFRAME
# ------------------------
def create_devices_dataframe(excel_file, excel_sheet_name):

    """
    PURPOSE
    -------
    Create a DataFrame from the data in the table of the argumented EXCEL file and sheet

    
    ARGUMENTS
    ---------
    excel_file       (str):            Absolute path to the Excel (.xlsx) file.
    excel_sheet_name (str):            Name of the worksheet to update.

    
    RETURN VALUE
    ------------
    If successful, it returns a DataFrame with all the table data. Otherwise, it returns an error code
    """
    
    try:
        # Read the data from the specified sheet
        all_devices_df = pd.read_excel(excel_file, sheet_name=excel_sheet_name)
        #-------------------------------------------
        logger.info("Dataframe creation succeeded!")
        #-------------------------------------------
    
    except FileNotFoundError:
        #------------------------------------------------------------------------------
        logger.error(f"Dataframe creation failed: The path {excel_file} was not found")
        #------------------------------------------------------------------------------
        return "FILE_NOT_FOUND_ERROR"
    
    except Exception as e:

        # Check if this is specifically a sheet name error
        if "Worksheet named" in str(e) and "not found" in str(e):
            #-----------------------------------------------------------------------------------------------------
            logger.error(f"Dataframe creation failed: Worksheet '{excel_sheet_name}' not found in the Excel file")
            #-----------------------------------------------------------------------------------------------------
            return "SHEET_NOT_FOUND_ERROR"
        else:
            #----------------------------------------------------------------------------------
            logger.error(f"Dataframe creation failed: error while reading the EXCEL file: {e}")
            #----------------------------------------------------------------------------------
            return "UNEXPECTED_ERROR"
    
    if 'OOBM IP Address' not in all_devices_df.columns:
        #-----------------------------------------------------------------------------------------------------------
        logger.error(f"Dataframe creation failed: a column named 'OOBM IP Address' was not found in the EXCEL file")
        #-----------------------------------------------------------------------------------------------------------
        return "COLUMN_NOT_FOUND_ERROR"
    
    # If successful, return the DataFrame
    return all_devices_df


# VALIDATE DEVICES DATAFRAME
# --------------------------
def valid_devices_dataframe(all_devices_df):

    """
    PURPOSE
    -------
    Validates that the DataFrame contains all required columns and wipes
    dynamic tracking columns to ensure a fresh state for each run.
    Remove any rows in the Dataframe that don't have an OOBM IP Address value,
    or a Hostname value, an invalid IP address, or a missing Model. It also sets 
    the current IOS version value for online devices and it points out the offline ones.

    
    ARGUMENTS
    ---------
    all_devices_df (str): the DataFrame with all the table data

    
    RETURN VALUE
    ------------
    Filtered DataFrame with rows that have valid IP Addresses only, both Online and Offline.
    """

    ### <=== REMOVE ROWS WITH MISSING CRITICAL VALUES ===> ###
    # We will only drop rows that don't have a valid value for any of the subset columns
    valid_devices_df = all_devices_df.dropna(subset=['Hostname', 'Model', 'OOBM IP Address']).copy()
    

    ### <=== HANDLE COLUMN INDEX ISSUES DUE TO MISSING COLUMNS ===> ###
    expected_columns = [
        'Current IOS Version', 'Needs Update', 'Update IOS File Present', 'Enough Flash Space',
        'Status', 'Auth Status', 'Transfer Result', 'Install Status', 'Update Result', 'Cleaned Inactive'
    ]
    
    # Handle situations where a column might be missing
    missing_columns = [col for col in expected_columns if col not in valid_devices_df.columns]
    if missing_columns:
        #------------------------------------------------------------------------------------------------
        logger.error(f"DataFrame validation failed: missing column(s) in Excel sheet: {missing_columns}")
        #------------------------------------------------------------------------------------------------
        return None
    
    
    ### <=== WIPE DYNAMIC COLUMNS FOR FRESH START ===> ### 
    # Dynamic columns that should be wiped on each run. Note that 'expected_columns' has already been validated to have
    # columns that are dynamic and also present on the DataFrame.
    for col in expected_columns:
        valid_devices_df[col] = ''


    ### <=== HANDLE VALUE TYPES FOR BLANK COLUMNS ===> ###
    # When pandas reads an Excel column that has no values in it (all cells empty), it infers the dtype as float64 — since NaN is a 
    # float in numpy, which is what pandas uses to represent empty cells. So when we later try to write a string like '17.13.1a' 
    # into that column with .at[index, ...], pandas refuses because it can't fit a string into a float64 column:
        
    #     TypeError: Invalid value '17.13.1a' for dtype 'float64'
    
    # Initialize all existent columns that will receive mixed string values — blank Excel columns are read as float64
    for col in expected_columns:
        valid_devices_df[col] = valid_devices_df[col].astype(object)


    ### <=== REMOVE ROWS WITH INVALID IP ADDRESS ===> ###
    # This line takes the 'OOBM IP Address' column and runs is_valid_ip() on each value individually. Then, it stores on the variable
    # the rows that contain an invalid IP only, so we can log them and handle errors better before removing them
    invalid_ips = valid_devices_df[valid_devices_df['OOBM IP Address'].apply(ip_addressing_ops.is_valid_ip) == False]

    # Log any rows that were dropped due to invalid IPs
    for _, row in invalid_ips.iterrows():
        #---------------------------------------------------------------------------------------------------------------
        logger.warning(f"Skipping device '{row['Hostname']}': invalid OOBM IP Address value '{row['OOBM IP Address']}'")
        #---------------------------------------------------------------------------------------------------------------

    # Filter out any rows with invalid IPs
    valid_devices_df = valid_devices_df[valid_devices_df['OOBM IP Address'].apply(ip_addressing_ops.is_valid_ip) == True]

    ### <=== RETURN FILTERED DATAFRAME ===> ###
    return valid_devices_df
  

# POPULATE THE STATUS AND AUTH STATUS COLUMN
# ------------------------------------------
def populate_status_and_auth_status_column(valid_devices_df, username, password):

    """
    PURPOSE
    -------
    Enables RESTCONF, SCP and HTTPS on all devices and populates the
    'Status' column with ONLINE or OFFLINE, and the 'Auth Status' column
    with AUTH_OK, AUTH_BAD, or None based on the result.

    
    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing at minimum 'OOBM IP Address' and 'Hostname' columns.
    username         (str):          Device SSH username.
    password         (str):          Device SSH password.

    
    RETURN VALUE
    ------------
    None if successful, or an error code string on failure:
        - "NO_DEVICES_ERROR"        : DataFrame contains no devices to process
        - "CONNECTION_TIMEOUT_ERROR": All devices failed to respond (timeout)
        - "AUTH_FAILED_ERROR"       : All devices failed authentication
        - "UNEXPECTED_ERROR"        : An unexpected exception occurred
    """

    try:
        # Check if there are any devices to process
        if valid_devices_df.empty:
            #-----------------------------------------------------------
            logger.error("No devices found in the DataFrame to process")
            #-----------------------------------------------------------
            return "NO_DEVICES_ERROR"
        
        results = device_cli_ops.enable_scp_and_restconf_all(valid_devices_df, username, password)
        
        # Check if we got any results back
        if not results:
            #---------------------------------------------------------
            logger.error("No results returned from device operations")
            #---------------------------------------------------------
            return "CONNECTION_TIMEOUT_ERROR"
        
        # Track statistics for error detection
        online_count = 0
        auth_ok_count = 0
        
        # Write Status and Auth Status back to the dataframe for each device
        for index, status, auth_status in results:
            valid_devices_df.at[index, 'Status']      = status
            valid_devices_df.at[index, 'Auth Status'] = auth_status
            
            if status == 'ONLINE':
                online_count += 1
            if auth_status == 'AUTH_OK':
                auth_ok_count += 1
        
        # If all devices are offline, return timeout error
        if online_count == 0 and len(results) > 0:
            #-----------------------------------------------------
            logger.error("All devices are offline or unreachable")
            #-----------------------------------------------------
            return "CONNECTION_TIMEOUT_ERROR"
        
        # If devices are online but none authenticated, return auth error
        if online_count > 0 and auth_ok_count == 0:
            #---------------------------------------------------------------------------
            logger.error("All online devices failed authentication — check credentials")
            #---------------------------------------------------------------------------
            return "AUTH_FAILED_ERROR"
        
        return None  # Success — at least some devices worked
        
    except Exception as e:
        return "UNEXPECTED_ERROR"


# POPULATE THE RESTCONF STATUS COLUMN 
# -----------------------------------
def populate_restconf_status_column(valid_devices_df, username, password, timeout, interval):

    """
    PURPOSE
    -------
    Polls the RESTCONF root endpoint on all ONLINE/AUTH_OK devices and populates
    the 'RESTCONF Status' column with OPERATIVE, NOT_OPERATIVE, or N/A. This column only
    exists in the DataFrame.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing at minimum 'OOBM IP Address', 'Status', and 'Auth Status' columns.
    username         (str):          Device username for RESTCONF authentication.
    password         (str):          Device password for RESTCONF authentication.
    timeout          (int):          Maximum seconds to wait for RESTCONF to become operative.
    interval         (int):          Seconds between polling attempts.


    RETURN VALUE
    ------------
    None if successful, or an error code string on failure:
        - "NO_ELIGIBLE_DEVICES_ERROR": No ONLINE/AUTH_OK devices to poll
        - "RESTCONF_TIMEOUT_ERROR"   : All eligible devices failed RESTCONF check
        - "UNEXPECTED_ERROR"         : An unexpected exception occurred
    """

    try:
        # Only poll devices that are reachable and authenticated — others can't have RESTCONF operative
        eligible_devices = valid_devices_df[
            (valid_devices_df['Status'] == 'ONLINE') &
            (valid_devices_df['Auth Status'] == 'AUTH_OK')
        ]

        # Check if there are any devices to poll
        if eligible_devices.empty:
            # Still populate N/A for all rows
            for index, row in valid_devices_df.iterrows():
                valid_devices_df.at[index, 'RESTCONF Status'] = 'N/A'
            return "NO_ELIGIBLE_DEVICES_ERROR"

        # Returns {ip: True/False} — True if RESTCONF became operative, False if timed out
        restconf_results = asyncio.run(device_api_ops.get_all_restconf_status(
            eligible_devices, username, password, timeout, interval
        ))

        # Track statistics
        operative_count = 0
        
        # Map results back to the DataFrame by IP address
        for index, row in valid_devices_df.iterrows():
            ip = row['OOBM IP Address']

            # Device was not polled — RESTCONF check is irrelevant
            if row['Status'] != 'ONLINE' or row['Auth Status'] != 'AUTH_OK':
                valid_devices_df.at[index, 'RESTCONF Status'] = 'N/A'
                continue

            # True → OPERATIVE, False → NOT_OPERATIVE
            is_operative = restconf_results.get(ip, False)
            valid_devices_df.at[index, 'RESTCONF Status'] = 'OPERATIVE' if is_operative else 'NOT_OPERATIVE'
            
            if is_operative:
                operative_count += 1

        # Log summary
        total_eligible = len(eligible_devices)
        logger.info(f"RESTCONF check complete: {operative_count}/{total_eligible} devices operative")

        # If no devices became operative, return timeout error
        if operative_count == 0 and total_eligible > 0:
            return "RESTCONF_TIMEOUT_ERROR"

        return None  # Success — at least some devices are RESTCONF operative

    except Exception as e:
        return "UNEXPECTED_ERROR"

        
# POPULATE CURRENT VERSION COLUMN
# -------------------------------
def populate_current_version_column(valid_devices_df, username, password):
    
    """
    PURPOSE
    -------
    Retrieves the current IOS version for all online devices and populates
    the 'Current IOS Version' column.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing at minimum 'OOBM IP Address' and 'RESTCONF Status' columns.
    username         (str):          Device username for RESTCONF authentication.
    password         (str):          Device password for RESTCONF authentication.


    RETURN VALUE
    ------------
    None if successful, or an error code string on failure:
        - "NO_OPERATIVE_DEVICES_ERROR": No devices with OPERATIVE RESTCONF status
        - "VERSION_RETRIEVAL_FAILED_ERROR": All operative devices failed version retrieval
        - "UNEXPECTED_ERROR": An unexpected exception occurred
    """

    try:
        # Run all the checks on online devices with operative RESTCONF only
        online_devices = valid_devices_df[valid_devices_df['RESTCONF Status'] == 'OPERATIVE']
        
        # Check if there are any devices to query
        if online_devices.empty:
            #------------------------------------------------------------------------------------------
            logger.error("No devices with OPERATIVE RESTCONF status available for version retrieval")
            #------------------------------------------------------------------------------------------
            # Still populate empty strings for all rows
            valid_devices_df['Current IOS Version'] = ''
            return "NO_OPERATIVE_DEVICES_ERROR"
        
        results = asyncio.run(device_api_ops.get_all_versions(online_devices, username, password))
        
        # Check if we got any results back
        if not results:
            valid_devices_df['Current IOS Version'] = ''
            return "VERSION_RETRIEVAL_FAILED_ERROR"
        
        # Map (ip, version) tuples back to the dataframe by IP address
        version_map = {ip: version if version else '' for ip, version in results}
        valid_devices_df['Current IOS Version'] = valid_devices_df['OOBM IP Address'].map(version_map).fillna('')
        
        # Track statistics
        successful_count = sum(1 for v in version_map.values() if v)
        total_queried = len(online_devices)
        
        
        # If no devices returned a version, return error
        if successful_count == 0 and total_queried > 0:
            return "VERSION_RETRIEVAL_FAILED_ERROR"
        
        return None  # Success — at least some devices returned a version
        
    except Exception as e:
        return "UNEXPECTED_ERROR"
        

# DETERMINE IF THE DEVICE NEEDS AN UPDATE
# ---------------------------------------
def populate_needs_update_column(valid_devices_df):

    """
    PURPOSE
    -------
    Compares the current IOS version against the recommended IOS version
    for each device, and stores the result in the 'Needs Update' column.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing 'Current IOS Version' and 'Recommended IOS Version' columns.


    RETURN VALUE
    ------------
    None, modifying valid_devices_df in place via .at[] assignment, or an error code string on failure:
        - "NO_DEVICES_NEED_UPDATE": All devices are up to date or have unknown status
        - "MISSING_VERSION_DATA_ERROR": No devices have both current and recommended versions
        - "UNEXPECTED_ERROR": An unexpected exception occurred
    """

    try:
        # Track statistics
        needs_update_count = 0
        up_to_date_count = 0
        unknown_count = 0
        missing_data_count = 0
        
        for index, row in valid_devices_df.iterrows():

            hostname = row['Hostname']
            current_version = row['Current IOS Version']
            recommended_version = row['Recommended IOS Version']

            # Skip rows where either value is missing
            if pd.isna(current_version) or pd.isna(recommended_version) or current_version == '' or recommended_version == '':
                valid_devices_df.at[index, 'Needs Update'] = 'UNKNOWN'
                missing_data_count += 1
                continue
            
            # Normalize the strings so the 0s in front of each digit don't make it look like they are different versions
            current_version_norm = common_helper_functions.normalize_ios_version_string(current_version)
            recommended_version_norm = common_helper_functions.normalize_ios_version_string(recommended_version)
            
            # Write NO if versions match, YES if they differ
            if current_version_norm == recommended_version_norm:
                valid_devices_df.at[index, 'Needs Update'] = 'NO'
                up_to_date_count += 1
                #--------------------------------------------------------------------
                logger.info(f"'{hostname}': up to date (current: {current_version})")
                #--------------------------------------------------------------------
            else:
                valid_devices_df.at[index, 'Needs Update'] = 'YES'
                needs_update_count += 1
                #-----------------------------------------------------------------------------------------------
                logger.info(f"'{hostname}': needs update (current: {current_version} to {recommended_version})")
                #-----------------------------------------------------------------------------------------------
        
        # Also mark any remaining rows that weren't processed (shouldn't happen, but safe)
        for index in valid_devices_df.index:
            if pd.isna(valid_devices_df.at[index, 'Needs Update']) or valid_devices_df.at[index, 'Needs Update'] == '':
                valid_devices_df.at[index, 'Needs Update'] = 'UNKNOWN'
                unknown_count += 1
        
        # Log summary
        total_devices = len(valid_devices_df)
        #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        logger.info(f"Update check complete: {needs_update_count} need update, {up_to_date_count} up to date, {missing_data_count + unknown_count} unknown out of {total_devices} total devices")
        #----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        # Check if any devices have version data at all
        if missing_data_count == total_devices:
            #------------------------------------------------------------------------
            logger.error("No devices have both current and recommended version data")
            #------------------------------------------------------------------------
            return "MISSING_VERSION_DATA_ERROR"
        
        # Check if any devices need an update
        if needs_update_count == 0:
            if up_to_date_count > 0:
                #-----------------------------------------------------------------------------------------
                logger.warning(f"All {up_to_date_count} devices with version data are already up to date")
                #-----------------------------------------------------------------------------------------
                return "NO_DEVICES_NEED_UPDATE"
            else:
                #------------------------------------------------------------------------
                logger.warning("No devices need an update (all unknown or missing data)")
                #------------------------------------------------------------------------
                return "NO_DEVICES_NEED_UPDATE"
        
        return None  # Success — at least one device needs an update
        
    except Exception as e:
        logger.error(f"Unexpected error during update check: {e}")
        return "UNEXPECTED_ERROR"


# GET THE RECOMMENDED IOS IMAGE FILE PATH
# ---------------------------------------
def valid_devices_df_with_image_path(valid_devices_df):

    """
    PURPOSE
    -------
    Locates the IOS image file for each device in the DataFrame by matching
    the recommended IOS version against files in the model-specific subfolder
    of the IOS image repository. Adds the absolute path of the matched image
    as a new 'IOS Image Path' column.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing at minimum 'Model' and 'Recommended IOS Version' columns.


    RETURN VALUE
    ------------
    None (if at least one device needing update has a valid image path) - modifies valid_devices_df in place via .at[] assignment. 
        - "IOS_REPOSITORY_NOT_FOUND_ERROR": The ios_repository folder doesn't exist
        - "NO_IMAGE_FOUND_ERROR": No matching IOS images found for any device needing update
        - "UNEXPECTED_ERROR": An unexpected exception occurred
    """

    try:
        # First, lets retrieve the absolute path of the folder that contains all the subfolders of the IOS images
        ios_image_repository_path = 'ios_repository'

        # Check if the repository folder exists
        if not os.path.isdir(ios_image_repository_path):
            #------------------------------------------------------------------------------
            logger.error(f"IOS image repository not found at: {ios_image_repository_path}")
            #------------------------------------------------------------------------------
            # Still populate columns for all rows
            for index, row in valid_devices_df.iterrows():
                if row['Needs Update'] == 'YES':
                    valid_devices_df.at[index, 'IOS Image Path'] = None # this column is not on the EXCEL file
                    valid_devices_df.at[index, 'Update IOS File Present'] = 'NO'
                else:
                    valid_devices_df.at[index, 'IOS Image Path'] = None # this column is not on the EXCEL file
                    valid_devices_df.at[index, 'Update IOS File Present'] = 'N/A'
            return "IOS_REPOSITORY_NOT_FOUND_ERROR"
        
        # Track statistics for devices needing update
        devices_needing_update = valid_devices_df[valid_devices_df['Needs Update'] == 'YES']
        total_needing_update = len(devices_needing_update)
        image_found_count = 0

        # Iterate over each DataFrame row
        for index, row in valid_devices_df.iterrows():

            # Device doesn't need an update (it can also be that it failed to retrieve current version, or missing recommended one) — image path lookup is irrelevant
            if row['Needs Update'] in ('NO', 'UNKNOWN'):
                valid_devices_df.at[index, 'IOS Image Path'] = None # keep things simple, don't use N/A
                valid_devices_df.at[index, 'Update IOS File Present'] = 'N/A' # we want to let the user know in the EXCEL
                continue

            model = row['Model'] # we will use this variable to locate the folder with the IOS images for such model
            recommended_version = row['Recommended IOS Version'] # we will use this variable to locate the exact IOS image within the model folder

            # Build the path to the model-specific subfolder
            model_folder_path = os.path.join(ios_image_repository_path, model)

            # Check the model folder exists
            if not os.path.isdir(model_folder_path):
                #---------------------------------------------------------------------------
                logger.warning(f"Model folder not found for '{model}': {model_folder_path}")
                #---------------------------------------------------------------------------
                # When this continue fires, execution jumps to the next iteration of the for index, row loop. The if ios_image_path / else block at the bottom 
                # never runs for this device — so 'IOS Image Path' and 'Update IOS File Present' are never set for it. Fix:
                valid_devices_df.at[index, 'IOS Image Path'] = None
                valid_devices_df.at[index, 'Update IOS File Present'] = 'NO'
                continue
            
            # Search for the IOS image file whose name contains the recommended version string
            ios_image_path = None
            for file in os.listdir(model_folder_path):
                # The version in the filename (17.15.04c) uses a different format than what RESTCONF returns (17.15.4c) — note the zero-padded 04 vs 4. 
                # So a direct string match won't always work. The safest approach is to normalize both strings before comparing — strip leading zeros 
                # from each numeric segment:
                if common_helper_functions.normalize_ios_version_string(recommended_version) in common_helper_functions.normalize_ios_version_string(file):
                    ios_image_path = os.path.join(model_folder_path, file)
                    break

            if ios_image_path:
                #----------------------------------------------------------------------------------------
                logger.info(f"IOS image found for '{model}' - '{recommended_version}': {ios_image_path}")
                #----------------------------------------------------------------------------------------
                valid_devices_df.at[index, 'IOS Image Path'] = ios_image_path
                valid_devices_df.at[index, 'Update IOS File Present'] = 'YES' 
                image_found_count = image_found_count + 1

            else:
                #------------------------------------------------------------------------------------------------------------------
                logger.warning(f"No IOS image found for '{model}' matching version '{recommended_version}' in {model_folder_path}")
                #------------------------------------------------------------------------------------------------------------------
                valid_devices_df.at[index, 'Update IOS File Present'] = 'NO'
                valid_devices_df.at[index, 'IOS Image Path'] = None

        # Outside the loop, otherwise only 1 device data will be filled!
        # If no images found for any device needing update, return error
        if total_needing_update > 0 and image_found_count == 0:
            #------------------------------------------------------------------------------------------------
            logger.error(f"No IOS images found for any of the {total_needing_update} devices needing update")
            #------------------------------------------------------------------------------------------------
            return "NO_IMAGE_FOUND_ERROR"
        
        # Outside the loop, otherwise only 1 device data will be filled!
        return None  # Success — at least one device has a valid image path
        
    except Exception as e:
        #--------------------------------------------------------------
        logger.error(f"Unexpected error during image path lookup: {e}")
        #--------------------------------------------------------------
        return "UNEXPECTED_ERROR"


# GET THE SIZE OF THE IOS IMAGE FILE
# ----------------------------------
def get_image_files_size(valid_devices_df):

    """
    PURPOSE
    -------
    Retrieves the file size in bytes of each IOS image file referenced
    in the DataFrame, for rows that have their "Needs Update" value set to YES,
    and stores the result in a new 'IOS Image Size' column. 
    Rows with a None value in 'IOS Image Path' are skipped.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing an 'IOS Image Path' column with absolute paths to the IOS image files.


    RETURN VALUE
    ------------
    None if at least one device needing update has a valid image size, or an error code string on failure. It also modifies valid_devices_df in place via .at[] assignment.
    """

    try:
        # Only retrieve image size for devices that need an update
        devices_needing_update = valid_devices_df[valid_devices_df['Needs Update'] == 'YES']
        
        # It does not make sense to handle an error of 'devices_needing_update' being empty. We already handle that error on
        # the 'populate_needs_update_column' function, so it would never been caught here!

        # Track statistics
        total_needing_update = len(devices_needing_update)
        valid_size_count = 0
        missing_path_count = 0
        file_not_found_count = 0
        
        # Note that we are using the original index values from valid_devices_df — iterrows() preserves the original index.
        for index, row in devices_needing_update.iterrows():

            image_path = row['IOS Image Path']
            hostname = row['Hostname']

            # We can safely skip rows where the image path was not resolved
            if pd.isna(image_path):
                #---------------------------------------------------------------------------
                logger.warning(f"Skipping '{row['Hostname']}': no IOS image path available")
                #---------------------------------------------------------------------------
                valid_devices_df.at[index, 'IOS Image Size'] = None
                missing_path_count += 1
                continue

            # Check the file actually exists before attempting to read its size
            if not os.path.isfile(image_path):
                #------------------------------------------------------------------------------
                logger.warning(f"Skipping '{row['Hostname']}': file not found at {image_path}")
                #------------------------------------------------------------------------------
                valid_devices_df.at[index, 'IOS Image Size'] = None
                file_not_found_count += 1
                continue

            # Get the file size in bytes
            try:
                image_size = os.path.getsize(image_path)
                #--------------------------------------------------------------------------------------------
                logger.info(f"Image size for '{hostname}': {image_size / (1024*1024):.0f} MB - {image_path}")
                #--------------------------------------------------------------------------------------------
                valid_devices_df.at[index, 'IOS Image Size'] = image_size
                valid_size_count += 1

            except OSError as e:
                #------------------------------------------------------------------------------
                logger.error(f"Failed to read file size for '{hostname}' at {image_path}: {e}")
                #------------------------------------------------------------------------------
                valid_devices_df.at[index, 'IOS Image Size'] = None
                file_not_found_count += 1

        # For devices not needing update, set size to None
        for index, row in valid_devices_df.iterrows():
            if row['Needs Update'] in ('NO', 'UNKNOWN'):
                valid_devices_df.at[index, 'IOS Image Size'] = None
        
        # If no valid image sizes retrieved for any device needing update, return error
        if total_needing_update > 0 and valid_size_count == 0:
            #---------------------------------------------------------------------------------------------------------------
            logger.error(f"No valid image files could be read for any of the {total_needing_update} devices needing update")
            #---------------------------------------------------------------------------------------------------------------
            return "NO_VALID_IMAGE_FILES_ERROR"
        
        return None  # Success — at least one device has a valid image size
        
    except Exception as e:
        #-----------------------------------------------------------------
        logger.error(f"Unexpected error during image size retrieval: {e}")
        #-----------------------------------------------------------------
        return "UNEXPECTED_ERROR"


# GET AND SET THE FLASH FREE SPACE
# --------------------------------
def populate_flash_free_space_column(valid_devices_df, username, password):
    """
    PURPOSE
    -------
    Runs the async flash free space retrieval, then maps the results
    back into the DataFrame as a new 'Flash Free Space' column.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing at minimum 'OOBM IP Address' column.
    username         (str):          Device username for RESTCONF authentication.
    password         (str):          Device password for RESTCONF authentication.


    RETURN VALUE
    ------------
    None - modifies valid_devices_df in place via .at[] assignment. 
    """

    # Only query flash space for devices that need an update AND are ONLINE with AUTH_OK
    devices_needing_update_authgood_and_online = valid_devices_df[
        (valid_devices_df['Needs Update'] == 'YES') &
        (valid_devices_df['Status'] == 'ONLINE') &
        (valid_devices_df['Auth Status'] == 'AUTH_OK')
    ]

    total_eligible = len(devices_needing_update_authgood_and_online)
    
    # Use your threaded function instead of asyncio
    results = device_cli_ops.get_flash_free_space_all(
        devices_needing_update_authgood_and_online,
        username,
        password
    )

    # We need the error counts for situations if all devices had this errors:
    error_counts = {
            "NO_FLASH_FOUND": 0,
            "PARSE_ERROR": 0,
            "THREAD_ERROR": 0,
            "UNEXPECTED_ERROR": 0
        }

    # Map (ip, free_bytes) tuples back to the dataframe. Then .map() looks up each row's OOBM IP Address value in that dictionary. 
    # So even if the async results came back in a completely random order, the correct value would still land on the correct row
    free_space_map = {}

    for ip, free_bytes, error_code in results:
        free_space_map[ip] = free_bytes
        if error_code and error_code in error_counts:
            error_counts[error_code] += 1
        elif error_code:
            error_counts["UNEXPECTED_ERROR"] += 1

    valid_devices_df['Flash Free Space'] = valid_devices_df['OOBM IP Address'].map(free_space_map)
    
    # NOTE: Devices that need an update get their value from free_space_map, and devices that don't need an update get NaN from .map() (since their IP won't be in the map)

    # Count successful retrievals. Remember that all devices that failed to parse the free space (or another error) returned 'None' as the 'free_bytes' value in 'device_cli_ops.get_flash_free_space_all'.
    successful_count = sum(1 for v in free_space_map.values() if v is not None)
    
    # If no devices returned valid flash space, analyze the failure pattern
    if successful_count == 0:
        total_errors = sum(error_counts.values())
        
        if error_counts["PARSE_ERROR"] + error_counts["NO_FLASH_FOUND"] == total_errors:
            #--------------------------------------------------------------------------------------
            logger.error(f"All {total_eligible} eligible devices failed due to CLI parsing issues")
            #--------------------------------------------------------------------------------------
            return "ALL_DEVICES_PARSE_ERROR"
        else:
            #----------------------------------------------------------------------------------
            logger.error(f"All {total_eligible} eligible devices failed flash space retrieval")
            #----------------------------------------------------------------------------------
            return "ALL_DEVICES_FAILED_ERROR"
    
    return None  # Success — at least one device returned flash space


# DETERMINE IF FLASH FREE SPACE IS ENOUGH
# ---------------------------------------
def populate_enough_space_column(valid_devices_df):

    """
    PURPOSE
    -------
    Compares the available flash free space against the IOS image file size
    for each device, and stores the result in an existent 'Enough Flash Space' column.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing 'Flash Free Space' and 'IOS Image Size' columns.


    RETURN VALUE
    ------------
    None - modifies valid_devices_df in place via .at[] assignment.
    """

    # Iterate over each device row
    for index, row in valid_devices_df.iterrows():

        # Device doesn't need an update — flash space check is irrelevant
        if row['Needs Update'] in ('NO', 'UNKNOWN'):
            valid_devices_df.at[index, 'Enough Flash Space'] = 'N/A'
            continue

        # Retrieve the flash free space and image size for this device
        free_space = row['Flash Free Space']
        image_size = row['IOS Image Size']

        # Skip rows where either value is missing — can't compare if data is incomplete
        if pd.isna(free_space) or pd.isna(image_size):
            valid_devices_df.at[index, 'Enough Flash Space'] = "UNKNOWN"
            continue
        
        # Write YES if free space exceeds image size, NO otherwise.
        valid_devices_df.at[index, 'Enough Flash Space'] = 'YES' if free_space > image_size else 'NO'


# GET ELIGIBLE DEVICES FOR UPDATE
# -------------------------------
def get_eligible_devices_df(valid_devices_df):

    """
    PURPOSE
    -------
    Filters the DataFrame to only include devices that are eligible for
    an IOS update — online, sufficient flash space, valid image path,
    and flagged as needing an update.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): DataFrame containing 'Status','Auth Status,
    'Enough Flash Space', 'IOS Image Path', 'SCP Enabled' and 'Needs Update' columns.


    RETURN VALUE
    ------------
    New DataFrame containing only eligible devices.
    """

    # Apply conditions for eligibility
    eligible_mask = (
        (valid_devices_df['Status'] == 'ONLINE')                &   # Device responded to SSH — reachable and alive
        (valid_devices_df['RESTCONF Status'] == 'OPERATIVE')    &   # Device responded to the test API call
        (valid_devices_df['Current IOS Version'].notna())       &   # Current IOS version was successfully retrieved from the device
        (valid_devices_df['Enough Flash Space'] == 'YES')       &   # Sufficient flash space available to store the image
        (valid_devices_df['IOS Image Path'].notna())            &   # A matching IOS image file was found in the local repository
        (valid_devices_df['IOS Image Size'].notna())            &   # Image file size was successfully read from disk
        (valid_devices_df['Needs Update'] == 'YES')                 # Current version differs from the recommended version
    )

    # Define the eligible devices out of the argumented DataFrame
    eligible_devices_df = valid_devices_df[eligible_mask].copy()

    #------------------------------------------------------------------------------------------------------------------------
    logger.info(f"{len(eligible_devices_df)} device(s) eligible for IOS update out of {len(valid_devices_df)} valid devices")
    #------------------------------------------------------------------------------------------------------------------------

    return eligible_devices_df


# FOR CLI VERSION ONLY: USER SELECTION OF ELIGIBLE DEVICES
# --------------------------------------------------------
def select_devices_for_update(eligible_devices_df):
    """
    PURPOSE
    -------
    Prompts the user to select which eligible devices to push the IOS update to,
    by entering their DataFrame index numbers. Returns a new DataFrame containing
    only the selected devices.


    ARGUMENTS
    ---------
    eligible_devices_df (pd.DataFrame): DataFrame containing only eligible devices.


    RETURN VALUE
    ------------
    New DataFrame containing only the user-selected devices.
    """

    # Show the user the eligible devices and their indexes (showing is done in the main_execution_program)
    print("\nEnter the indexes of the devices you want to update, separated by commas (e.g. 0,1,3):")
    raw_input = input("> ").strip()

    try:
        # Parse the input into a list of integers
        selected_indexes = [int(i.strip()) for i in raw_input.split(',')]
    except ValueError:
        #--------------------------------------------------------------------------------
        logger.error(f"Invalid input: '{raw_input}' — expected comma-separated integers")
        #--------------------------------------------------------------------------------
        return None

    # Filter to only the rows whose index matches the user selection
    selected_eligible_devices_df = eligible_devices_df.loc[
        eligible_devices_df.index.isin(selected_indexes)
    ].copy()

    if selected_eligible_devices_df.empty:
        #-------------------------------------------------------------------------
        logger.error(f"No matching devices found for indexes: {selected_indexes}")
        #-------------------------------------------------------------------------
        return None

    #----------------------------------------------------------------------------------------------------
    logger.info(f"{len(selected_eligible_devices_df)} device(s) selected for update: {selected_indexes}")
    #----------------------------------------------------------------------------------------------------

    return selected_eligible_devices_df


# POPULATE THE TRANSFER STATUS COLUMN
# ------------------------------------
def populate_transfer_status_column(valid_devices_df, selected_eligible_devices_df, username, password, mode="threaded"):

    """
    PURPOSE
    -------
    Calls transfer_ios_image() for each selected eligible device and populates
    the 'Transfer Result' column in valid_devices_df with the return code.
    Prompts the user to choose between sequential and threaded transfer modes
    before starting. Devices not selected for transfer are marked as 'NOT_ATTEMPTED'.


    ARGUMENTS
    ---------
    valid_devices_df              (pd.DataFrame): Full device DataFrame — receives the Transfer Status values.
    selected_eligible_devices_df  (pd.DataFrame): Filtered DataFrame containing only devices selected for update.
    username                      (str):          Device SSH username.
    password                      (str):          Device SSH password.
    mode                          (str):          Threaded or Sequential. The default value is so the CLI version can work too.


    RETURN VALUE
    ------------
    None if at least one transfer succeeded, or an error code string:
        - "ALL_TRANSFERS_FAILED_ERROR" : All selected devices failed the file transfer
        - "UNEXPECTED_ERROR"           : An unexpected exception occurred
    """

    # Track success count
    success_count = 0
    total_selected = len(selected_eligible_devices_df)


    ### <=== MODE SELECTION ===> ###
    # Before starting any transfer, ask the user to choose how transfers will be executed.
    # Sequential is the safe default — one device at a time with a clean progress bar per device.
    # Threaded fires all transfers simultaneously, which is faster at scale but means multiple
    # progress bars will print concurrently and overwrite each other in the terminal.

    # FOR CLI Version
    #if mode == 'threaded':
       #threaded = True
        #print("\n  Threaded mode selected (GUI default).\n")
    #elif mode == 'sequential':
        #threaded = False
        #print("\n  Sequential mode selected.\n")
    # else:
    #     # Original interactive prompt (for CLI)
    #     print("\nSelect transfer mode:")
    #     print("  [1] Sequential — one device at a time, live progress bar per device. Slower but safer.")
    #     print("  [2] Threaded   — all devices concurrently. Faster but progress output will be interleaved.")
    #     print()

    #     while True:
    #         choice = input("Enter 1 or 2: ").strip()
    #         if choice in ('1', '2'):
    #             break
    #         print("Invalid input — please enter 1 or 2.")

    #     threaded = (choice == '2')

    #     if threaded:
    #         print("\n  WARNING: Threaded mode transfers to all devices simultaneously.")
    #         print("  Progress bars from multiple devices will overlap in the terminal.")
    #         print("  Check the Transfer Status column in the Excel tracker for individual results.\n")
    #     else:
    #         print("\n  Sequential mode selected — devices will be transferred one at a time.\n")

    try:

        ### <=== MARK NON-SELECTED DEVICES ===> ###
        # Every device in valid_devices_df that was not part of this transfer run gets
        # NOT_ATTEMPTED — this covers OFFLINE devices, NOT_OPERATIVE devices, devices with
        # insufficient flash space, and eligible devices the user chose not to select.
        # This ensures the Transfer Status column is fully populated for the Excel tracker.
        for index in valid_devices_df.index:
            if index not in selected_eligible_devices_df.index:
                valid_devices_df.at[index, 'Transfer Result'] = 'N/A'


        ### <=== TRANSFER LOGIC ===> ###
        # Nested helper shared by both execution modes. Extracts the parameters needed by
        # transfer_ios_image() from the DataFrame row, calls it, and returns the (index, result)
        # tuple so the caller can write the result back to the correct DataFrame row regardless
        # of execution order — which matters in threaded mode since futures complete out of order.
        def run_transfer(row):
            ip               = row['OOBM IP Address']
            hostname         = row['Hostname']
            local_image_path = row['IOS Image Path']

            # Derive the remote filename from the local path — the file lands on bootflash
            # with the same name it has in the local IOS image repository
            remote_filename  = os.path.basename(local_image_path)

            #-------------------------------------------------------
            logger.info(f"Starting transfer to '{hostname}' ({ip})")
            #-------------------------------------------------------

            result = device_file_transfer_ops.single_transfer_ios_image(
                ip,
                username,
                password,
                local_image_path,
                remote_filename
            )

            #-----------------------------------------------------------------------
            logger.info(f"Transfer Status for '{hostname}' ({ip}) set to: {result}")
            #-----------------------------------------------------------------------

            # Return the original DataFrame index alongside the result so the caller
            # can write back to the correct row in valid_devices_df
            return row.name, result


        ### <=== SEQUENTIAL ===> ###
        # Iterate over selected devices one at a time. Each transfer completes fully
        # before the next one starts — progress bar output is clean and readable.
        if mode == "sequential":
            #----------------------------------------------------
            logger.info("User elected a sequential SCP transfer")
            #----------------------------------------------------
            for _, row in selected_eligible_devices_df.iterrows():
                index, result = run_transfer(row)
                valid_devices_df.at[index, 'Transfer Result'] = result
                if result == 'SUCCESS':
                    success_count += 1


        ### <=== THREADED ===> ###
        elif mode == "threaded":
            #--------------------------------------------------
            logger.info("User elected a threaded SCP transfer")
            #--------------------------------------------------
            results = device_file_transfer_ops.threaded_transfer_ios_image_all(
                selected_eligible_devices_df, username, password
            )

            # results is {index: result} — write each back to the correct row
            for index, result in results.items():
                valid_devices_df.at[index, 'Transfer Result'] = result
                if result == 'SUCCESS':
                    success_count += 1


        # If all selected devices failed, return error
        if success_count == 0 and total_selected > 0:
            #--------------------------------------------------------------------------
            logger.error(f"All {total_selected} selected devices failed file transfer")
            #--------------------------------------------------------------------------
            return "ALL_TRANSFERS_FAILED_ERROR"


        return None
    

    except Exception as e:
        #-----------------------------------------------------
        logger.error(f"Unexpected error during transfer: {e}")
        #-----------------------------------------------------
        return "UNEXPECTED_ERROR"


# GET ELIGIBLE DEVICES FOR INSTALL
# --------------------------------
def get_install_eligible_devices_df(valid_devices_df):

    """
    PURPOSE
    -------
    Filters valid_devices_df to only include devices whose Transfer Result
    is SUCCESS — these are the only devices for which an install is applicable.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): Full device DataFrame.
        

    RETURN VALUE
    ------------
    New DataFrame containing only devices with Transfer Status SUCCESS.
    """

    install_eligible_devices_df = valid_devices_df[
        valid_devices_df['Transfer Result'] == 'SUCCESS'
    ].copy()

    #-----------------------------------------------------------------------------------------------------------------------------
    logger.info(f"{len(install_eligible_devices_df)} device(s) eligible for install out of {len(valid_devices_df)} valid devices")
    #-----------------------------------------------------------------------------------------------------------------------------

    return install_eligible_devices_df


# POPULATE THE INSTALL STATUS COLUMN
# ----------------------------------
# Confirm installs before reloading, per device (called by 'populate_install_status_column')
def confirm_installs(install_eligible_devices_df):

    """
    PURPOSE
    -------
    Presents each eligible device to the user one at a time and collects
    explicit YES/NO confirmation before any reload is triggered. Returns
    two separate DataFrames — one for confirmed devices and one for aborted
    ones — so the caller can pass only confirmed devices to install_ios_image_all
    and still record ABORTED in the results for the rest.


    ARGUMENTS
    ---------
    install_eligible_devices_df (pd.DataFrame): Devices eligible for install.


    RETURN VALUE
    ------------
    Tuple of (confirmed_df, aborted_df) — both are subsets of install_eligible_devices_df.
    """

    confirmed_indexes = []
    aborted_indexes   = []

    for idx, row in install_eligible_devices_df.iterrows():

        hostname = row['Hostname']
        ip       = row['OOBM IP Address']
        bin_file = os.path.basename(row['IOS Image Path'])

        print(f"\n  {'='*60}")
        print(f"  WARNING: About to run install on '{hostname}' ({ip})")
        print(f"  Command : install add file bootflash:{bin_file} activate commit")
        print(f"  Effect  : Device will reload. Session will drop.")
        print(f"  {'='*60}")

        while True:
            confirm = input(f"\n  Type 'YES' to proceed with install on '{hostname}', or 'NO' to abort: ").strip().upper()
            if confirm in ('YES', 'NO'):
                break
            print("  Invalid input — please type YES or NO.")

        if confirm == 'YES':
            confirmed_indexes.append(idx)
            #-------------------------------------------------------------
            logger.info(f"'{hostname}' ({ip}): install confirmed by user")
            #-------------------------------------------------------------
        else:
            aborted_indexes.append(idx)
            #--------------------------------------------------------------
            logger.warning(f"'{hostname}' ({ip}): install aborted by user")
            #--------------------------------------------------------------

    confirmed_df = install_eligible_devices_df.loc[confirmed_indexes].copy()
    aborted_df   = install_eligible_devices_df.loc[aborted_indexes].copy()

    #---------------------------------------------------------------------------------------------
    logger.info(f"{len(confirmed_df)} device(s) confirmed for install, {len(aborted_df)} aborted")
    #---------------------------------------------------------------------------------------------

    return confirmed_df, aborted_df

def populate_install_status_column(valid_devices_df, install_eligible_devices_df, username, password, skip_confirmation=False):

    """
    PURPOSE
    -------
    Orchestrates the full install sequence for all eligible devices:

        1. Populates the 'Install Status' column in valid_devices_df with the result for every device — confirmed, aborted, and not attempted.

        2. Collects per-device confirmation from the user before any reload is triggered (if skip_confirmation is False).

        3. Fires the install command concurrently on all confirmed devices.


    ARGUMENTS
    ---------
    valid_devices_df            (pd.DataFrame): Full device DataFrame — receives Install Status values.
    install_eligible_devices_df (pd.DataFrame): Devices eligible for install — Transfer Status
                                                SUCCESS and all eligibility conditions met.
    username                    (str):          Device SSH username.
    password                    (str):          Device SSH password.
    skip_confirmation           (bool):         When True (GUI mode), auto-confirms all devices without user input.


    RETURN VALUE
    ------------
    None if at least one device triggered the install, or an error code string on failure:
        - "ALL_INSTALLS_FAILED_ERROR": No device reached INSTALL_TRIGGERED (all ADD_FAILED, ACTIVATE_FAILED, CONNECT_ERROR, etc.)
        - "UNEXPECTED_ERROR"         : An unexpected exception occurred
    """

    try:

        ### <=== MARK NON-ELIGIBLE DEVICES ===> ###
        # Every device that was not part of this install run gets NOT_ATTEMPTED —
        # covers OFFLINE, NOT_OPERATIVE, insufficient flash, transfer failures,
        # and devices that never reached the eligibility check.
        for index in valid_devices_df.index:
            if index not in install_eligible_devices_df.index:
                valid_devices_df.at[index, 'Install Status'] = 'NOT_ATTEMPTED'


        ### <=== COLLECT CONFIRMATIONS ===> ###
        # All user confirmations are gathered upfront — before any thread is spawned —
        # so that install prompts never interleave with each other in the terminal.
        if skip_confirmation:
            confirmed_df = install_eligible_devices_df.copy()
            aborted_df = pd.DataFrame()
        else:
            confirmed_df, aborted_df = confirm_installs(install_eligible_devices_df)


        ### <=== RECORD ABORTED DEVICES ===> ###
        # Write ABORTED immediately for devices the user declined — these never reach
        # the install function so their status would otherwise remain unpopulated.
        # for idx in aborted_df.index:
        #     valid_devices_df.at[idx, 'Install Status'] = 'ABORTED'
            

        ### <=== FIRE INSTALLS ON CONFIRMED DEVICES ===> ###
        # Pass only confirmed devices to the threaded install function.
        # Returns {index: result} — one entry per device.
        triggered_count = 0
        if not confirmed_df.empty:
            install_results = device_cli_ops.install_ios_image_all(confirmed_df, username, password)

            for index, result in install_results.items():
                # Devices skipped due to failed transfer get N/A — install was never applicable
                valid_devices_df.at[index, 'Install Status'] = result
                if result == 'INSTALL_TRIGGERED':
                    triggered_count += 1
        else:
            #--------------------------------------------------------------------------
            logger.warning("No devices confirmed for install — skipping install phase")
            #--------------------------------------------------------------------------
            return "ALL_INSTALLS_FAILED_ERROR"
        
        # If no device triggered the install, it's a fatal error for the update
        if triggered_count == 0 and not confirmed_df.empty:
            #-----------------------------------------------------------------------------------
            logger.error(f"All {len(confirmed_df)} confirmed devices failed to trigger install")
            #-----------------------------------------------------------------------------------
            return "ALL_INSTALLS_FAILED_ERROR"
        
        return None  # At least one device is proceeding
    
    except Exception as e:
        #----------------------------------------------------------------------
        logger.error(f"Unexpected error during install status population: {e}")
        #----------------------------------------------------------------------
        return "UNEXPECTED_ERROR"


# POPULATE INSTALLED VERSION AND UPDATE RESULT COLUMNS
# -----------------------------------------------------
def populate_post_install_columns(valid_devices_df, install_eligible_devices_df, username, password):

    """
    PURPOSE
    -------
    Executes the post-reboot sequence for all devices that had an install
    triggered:

        1. Marks devices not part of the install run as N/A immediately.
        2. Waits 5 minutes for devices to complete their reload.
        3. Polls RESTCONF on all INSTALL_TRIGGERED devices for up to 5 minutes
           to determine if they came back online.
        4. Marks timed-out devices as UNKNOWN for both columns.
        5. Runs 'install commit' concurrently on all online devices.
        6. Retrieves the running version via RESTCONF and compares against
           the recommended version, populating both columns accordingly.


    ARGUMENTS
    ---------
    valid_devices_df            (pd.DataFrame): Full device DataFrame — receives column values.
    install_eligible_devices_df (pd.DataFrame): Devices that were part of the install run.
    username                    (str):          Device username.
    password                    (str):          Device password.


    RETURN VALUE
    ------------
    None if at least one triggered device came back online and processing completed,
    or an error code string on failure:
        - "ALL_DEVICES_TIMEOUT_ERROR": All triggered devices timed out during reload
        - "UNEXPECTED_ERROR"         : An unexpected exception occurred
    """

    try:

        ### <=== MARK NON-INSTALL DEVICES IMMEDIATELY ===> ###
        # Devices that were never part of the install run get N/A for both columns
        # upfront — no need to wait for them.
        for index in valid_devices_df.index:
            if index not in install_eligible_devices_df.index:
                valid_devices_df.at[index, 'Installed Version'] = 'N/A'
                valid_devices_df.at[index, 'Update Result']     = 'N/A'


        ### <=== FILTER TO INSTALL_TRIGGERED DEVICES ONLY ===> ###
        # Only devices that actually had the install command sent are relevant —
        # aborted devices are in install_eligible_devices_df but never reloaded
        # so they should not be polled.
        triggered_devices = valid_devices_df[
            valid_devices_df['Install Status'] == 'INSTALL_TRIGGERED'
        ]

        if triggered_devices.empty:
            #---------------------------------------------------------------------------
            logger.warning("No INSTALL_TRIGGERED devices found — skipping post-install")
            #---------------------------------------------------------------------------
            return None


        ### <=== INITIAL SLEEP — WAIT FOR RELOAD TO COMPLETE ===> ###
        # Give devices 5 minutes to complete the reload before starting to poll.
        # Install mode upgrades can take several minutes to finish booting. This
        # timer starts when the last device is marked with "INSTALL_TRIGGERED"
        #-----------------------------------------------------------------
        logger.info("Waiting 6 minutes for devices to complete reload...")
        #-----------------------------------------------------------------
        time.sleep(360)


        ### <=== POLL RESTCONF TO CHECK IF DEVICES ARE BACK ONLINE ===> ###
        # Use the existing wait_for_restconf() with a 10 minute timeout and 15 second
        # interval — returns {ip: True/False} for each triggered device.
        #-------------------------------------------------------------------------------------------
        logger.info("Polling RESTCONF to check if devices are back online (timeout: 10 minutes)...")
        #-------------------------------------------------------------------------------------------

        restconf_results = asyncio.run(
            device_api_ops.get_all_restconf_status(triggered_devices, username, password, 600, 15)
        )

        # Split triggered devices into those that came back up and those that timed out
        online_indexes  = []
        timeout_indexes = []

        for index, row in triggered_devices.iterrows():
            ip = row['OOBM IP Address']
            if restconf_results.get(ip):
                online_indexes.append(index)
            else:
                timeout_indexes.append(index)
                #------------------------------------------------------------------------------
                logger.error(f"'{row['Hostname']}' ({ip}): did not come back online — timeout")
                #------------------------------------------------------------------------------

        # Devices that timed out get UNKNOWN for both columns
        for index in timeout_indexes:
            valid_devices_df.at[index, 'Installed Version'] = 'UNKNOWN'
            valid_devices_df.at[index, 'Update Result']     = 'UNKNOWN'

        # If NO device came back online, it's a fatal error
        if not online_indexes:
            #-----------------------------------------------------------------------------------------
            logger.error(f"All {len(triggered_devices)} triggered devices failed to come back online")
            #-----------------------------------------------------------------------------------------
            return "ALL_DEVICES_TIMEOUT_ERROR"
        

        ### <=== RUN INSTALL COMMIT ON ALL ONLINE DEVICES ===> ###
        online_devices_df = triggered_devices.loc[online_indexes]

        if online_devices_df.empty:
            #--------------------------------------------------------------------------------
            logger.warning("No devices came back online — skipping commit and version check")
            #--------------------------------------------------------------------------------
            return

        # Returns {index: result} — one entry per device
        commit_results = device_cli_ops.commit_ios_install_all(online_devices_df, username, password)


        ### <=== RETRIEVE RUNNING VERSION AND POPULATE COLUMNS ===> ###
        version_results = asyncio.run(device_api_ops.get_all_versions(online_devices_df, username, password))
        version_map     = {ip: version for ip, version in version_results}

        for index, row in online_devices_df.iterrows():

            ip                  = row['OOBM IP Address']
            hostname            = row['Hostname']
            installed_version   = version_map.get(ip)
            recommended_version = row['Recommended IOS Version']
            commit_result       = commit_results.get(index, 'UNEXPECTED_ERROR')

            # Version retrieval failed — could not confirm running version
            if not installed_version:
                valid_devices_df.at[index, 'Installed Version'] = 'UNKNOWN'
                valid_devices_df.at[index, 'Update Result']     = 'UNKNOWN'
                #--------------------------------------------------------------------------
                logger.error(f"'{hostname}' ({ip}): version retrieval failed after commit")
                #--------------------------------------------------------------------------
                continue

            valid_devices_df.at[index, 'Installed Version'] = installed_version

            # Normalize both version strings before comparing to handle zero-padded
            # segments — '17.15.04c' and '17.15.4c' must be treated as equal
            installed_normalized   = common_helper_functions.normalize_ios_version_string(installed_version)
            recommended_normalized = common_helper_functions.normalize_ios_version_string(recommended_version)

            if installed_normalized != recommended_normalized:
                #----------------------------------------------------------------------------------------------------------------
                logger.warning(f"'{hostname}': version mismatch — installed {installed_version}, expected {recommended_version}")
                #----------------------------------------------------------------------------------------------------------------
                valid_devices_df.at[index, 'Update Result'] = 'FAILED'

            elif commit_result != 'COMMIT_SUCCESS':
                #-----------------------------------------------------------------------------------
                logger.warning(f"'{hostname}': version matches but commit failed — {commit_result}")
                #-----------------------------------------------------------------------------------
                valid_devices_df.at[index, 'Update Result'] = 'COMMIT_FAILED'

            else:
                #----------------------------------------------------------------------------------------------
                logger.info(f"'{hostname}': update successful — running {installed_version}, commit confirmed")
                #----------------------------------------------------------------------------------------------
                valid_devices_df.at[index, 'Update Result'] = 'SUCCESS'

        return None

    except Exception as e:
        #---------------------------------------------------------
        logger.error(f"Unexpected error during post-install: {e}")
        #---------------------------------------------------------
        return "UNEXPECTED_ERROR"


# POPULATE THE CLEANED INACTIVE COLUMN
# -------------------------------------
def populate_cleaned_inactive_column(valid_devices_df, username, password):

    """
    PURPOSE
    -------
    Runs 'install remove inactive' concurrently on all devices whose Install
    Status is INSTALL_TRIGGERED and populates the 'Cleaned Inactive' column
    in valid_devices_df with the return code. Devices not eligible for cleanup
    are marked as N/A.


    ARGUMENTS
    ---------
    valid_devices_df (pd.DataFrame): Full device DataFrame — receives Cleaned Inactive values.
    username         (str):          Device SSH username.
    password         (str):          Device SSH password.


    RETURN VALUE
    ------------
    None if cleanup succeeded for at least one triggered device (or no devices needed cleanup),
    or an error code string on failure:
        - "ALL_CLEANUP_FAILED_ERROR": All triggered devices failed cleanup (no CLEANED result)
        - "UNEXPECTED_ERROR"        : An unexpected exception occurred
    """

    try:
        # Returns {index: result} — N/A for non-triggered devices, return code for triggered ones
        results = device_cli_ops.remove_inactive_ios_all(valid_devices_df, username, password)

        # Count how many triggered devices actually succeeded in cleaning
        cleaned_count = 0
        triggered_count = 0
        for index, result in results.items():
            valid_devices_df.at[index, 'Cleaned Inactive'] = result
            # Result could be N/A (non-triggered), but we only care about triggered devices
            if result != 'N/A':
                triggered_count += 1
                if result == 'CLEANED':
                    cleaned_count += 1

        # If there were triggered devices but none cleaned successfully, it's a failure
        if triggered_count > 0 and cleaned_count == 0:
            #--------------------------------------------------------------------------
            logger.error(f"Cleanup failed for all {triggered_count} triggered devices")
            #--------------------------------------------------------------------------
            return "ALL_CLEANUP_FAILED_ERROR"

        return None   # Success or no triggered devices

    except Exception as e:
        #----------------------------------------------------
        logger.error(f"Unexpected error during cleanup: {e}")
        #----------------------------------------------------
        return "UNEXPECTED_ERROR"


# UPDATE THE EXCEL TRACKER
# ------------------------
def update_excel_tracker(excel_file, excel_sheet_name, valid_devices_df):

    """
    PURPOSE
    -------
    Opens an existing Excel workbook and updates the columns that host dynamic data
    for each device whose hostname matches an entry in the provided DataFrame.

    
    ARGUMENTS
    ---------
    excel_file       (str):            Absolute path to the Excel (.xlsx) file.
    excel_sheet_name (str):            Name of the worksheet to update.
    valid_devices_df (pd.DataFrame): DataFrame containing pertinent columns

    
    RETURN VALUE
    ------------
    If successful, it returns a success code. Otherwise, it returns an error code

    """

    # Initialize the variable
    workbook = None

    try:
        # Open the existing workbook and select the target worksheet
        workbook = load_workbook(excel_file) 
        worksheet = workbook[excel_sheet_name]

        # Find column indexes for all dynamic columns
        hostname_col_index = -1
        current_ios_col_index = -1
        status_col_index = -1
        auth_status_col_index = -1
        enough_flash_space_col_index = -1
        needs_update_col_index = -1
        update_ios_file_present_col_index = -1
        transfer_status_col_index = -1
        install_status_col_index = -1
        update_result_col_index = -1
        cleaned_inactive_col_index = -1

        # Scan header row (row 1) to locate the relevant column indexes by name
        for column in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=column).value
            if header == "Hostname":
                hostname_col_index = column
            elif header == "Current IOS Version":
                current_ios_col_index = column
            elif header == "Status":
                status_col_index = column
            elif header == "Auth Status":
                auth_status_col_index  = column
            elif header == "Enough Flash Space":
                enough_flash_space_col_index = column
            elif header == "Needs Update":
                needs_update_col_index = column
            elif header == "Update IOS File Present":
                update_ios_file_present_col_index = column
            elif header == "Transfer Result":
                transfer_status_col_index = column
            elif header == "Install Status":
                install_status_col_index = column
            elif header == "Update Result":
                 update_result_col_index = column
            elif header == "Cleaned Inactive":
                cleaned_inactive_col_index = column

        # Build lookup dicts from the dataframe: {hostname: currrent_version}, ...
        version_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Current IOS Version']))
        status_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Status']))
        auth_status_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Auth Status']))
        enough_flash_space_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Enough Flash Space']))
        needs_update_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Needs Update']))
        update_ios_file_present_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Update IOS File Present']))
        transfer_status_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Transfer Result']))
        install_status_col_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Install Status']))
        update_result_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Update Result']))
        cleaned_inactive_lookup = dict(zip(valid_devices_df['Hostname'], valid_devices_df['Cleaned Inactive']))

        # After the for loop that scans headers, before the for loop that writes rows, lets check if any columns are missing:
        required_columns = {
            'Hostname':                hostname_col_index,
            'Current IOS Version':     current_ios_col_index,
            'Status':                  status_col_index,
            'Auth Status':             auth_status_col_index,
            'Enough Flash Space':      enough_flash_space_col_index,
            'Needs Update':            needs_update_col_index,
            'Update IOS File Present': update_ios_file_present_col_index,
            'Transfer Result':         transfer_status_col_index,
            'Install Status':          install_status_col_index,
            'Update Result':           update_result_col_index,
            'Cleaned Inactive':        cleaned_inactive_col_index,
        }

        missing = [name for name, idx in required_columns.items() if idx == -1]
        if missing:
            #--------------------------------------------------------------------------------
            logger.error(f"Missing column(s) in Excel sheet '{excel_sheet_name}': {missing}")
            #--------------------------------------------------------------------------------
            return "MISSING_COLUMN_EXCEL_ERROR"

        # Iterate over Excel rows and update only matching hostnames
        for row in range(2, worksheet.max_row + 1):
            hostname = worksheet.cell(row=row, column=hostname_col_index).value
            # If the hostname exists in the dataframe, overwrite all dynamic columns
            if hostname in version_lookup:
                worksheet.cell(row=row, column=current_ios_col_index, value=version_lookup[hostname])
                worksheet.cell(row=row, column=status_col_index, value=status_lookup[hostname])
                worksheet.cell(row=row, column=auth_status_col_index, value=auth_status_lookup[hostname])
                worksheet.cell(row=row, column=enough_flash_space_col_index, value=enough_flash_space_lookup[hostname])
                worksheet.cell(row=row, column=needs_update_col_index, value=needs_update_lookup[hostname])
                worksheet.cell(row=row, column=update_ios_file_present_col_index, value=update_ios_file_present_lookup[hostname])
                worksheet.cell(row=row, column=transfer_status_col_index, value=transfer_status_lookup[hostname])
                worksheet.cell(row=row, column=install_status_col_index, value=install_status_col_lookup[hostname])
                worksheet.cell(row=row, column=update_result_col_index, value=update_result_lookup[hostname])
                worksheet.cell(row=row, column=cleaned_inactive_col_index, value=cleaned_inactive_lookup[hostname])

        # Save the workbook back to disk, persisting all cell changes
        workbook.save(excel_file)
        #-------------------------------------------
        logger.info(f"Updating EXCEL file succeded")
        #-------------------------------------------
        return "SUCCESS"
        
    except Exception as e:

        if isinstance(e, PermissionError):
            #-----------------------------------------------------------------------------------------------------
            logger.error(f"Updating EXCEL file failed: Could not save the file, please ensure the file is closed")
            #-----------------------------------------------------------------------------------------------------
            return "PERMISSION_DENIED_ERROR"
        
        else:
            #-----------------------------------------------------------------------------
            logger.error(f"Updating EXCEL file failed: an unexpected error occured - {e}")
            #-----------------------------------------------------------------------------
            return "UNEXPECTED_ERROR"
    
    finally:
        if workbook:
            workbook.close()
            #------------------------------------------------------------
            logger.info(f"The EXCEL file has been closed by the program")
            #------------------------------------------------------------


# CHECK IF EXCEL FILE IS OPEN
# ---------------------------
def check_excel_file_not_open(excel_file):
    """
    PURPOSE
    -------
    Checks if the Excel file is currently open by another process.
    If it is, logs an error and returns an error code.


    ARGUMENTS
    ---------
    excel_file (str): Absolute path to the Excel (.xlsx) file.


    RETURN VALUE
    ------------
    None if the file is not open, or 'ERROR_EXCEL_FILE_OPEN' if it is locked by another process.
    """

    # Iterate over all running processes, retrieving their PID, name, and open file handles
    for process in psutil.process_iter(['pid', 'name', 'open_files']):

        try:
            # Iterate over the files currently open by this process (default to empty list if None)
            for open_file in process.info['open_files'] or []:
                # Case-insensitive check to see if the Excel file path matches any open file
                if excel_file.lower() in open_file.path.lower():
                    #---------------------------------------------------------------------------------------------------------------------------------------------
                    logger.error(f"EXCEL file is currently open by '{process.info['name']}' (PID {process.info['pid']}) - please close it and re-run the program")
                    #---------------------------------------------------------------------------------------------------------------------------------------------
                    # Stop the program — file must be closed before proceeding
                    return 'ERROR_EXCEL_FILE_OPEN'

        except (psutil.NoSuchProcess, psutil.AccessDenied):
            # Process may have died mid-iteration or denied access to its file list — skip it
            continue